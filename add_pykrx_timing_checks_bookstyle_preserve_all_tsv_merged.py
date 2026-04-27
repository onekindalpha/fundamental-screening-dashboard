#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
국장 린치/그레이엄 xlsx에 책 스타일 타이밍 자동체크 시트를 추가한다.

핵심 규칙
- 시장이 상승세인지 먼저 확인
- 업종(현재 시트 내 그룹 바스켓 기준)이 상승세인지 확인
- 종목은 MA30 위 / MA30 우상향 / 저항 돌파 / 거래량 급증 / 늦은 진입 아님 여부 확인
- 저항선은 최근 스윙 고점(피벗 고점) 전체를 사용

주의
- 업종상승세는 KRX 공식 업종지수 대신 "현재 시트 안 동일 그룹 종목 바스켓" 기준이다.
- 시장상승세는 KOSPI/KOSDAQ 벤치마크 기준이다.
"""

from __future__ import annotations

import argparse
import math
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def norm_code(x) -> str:
    s = str(x).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = ''.join(ch for ch in s if ch.isdigit())
    return s.zfill(6) if s else ''


def find_tsv_header_row(path: Path) -> int:
    preview = pd.read_csv(path, sep='\t', header=None, dtype=object, nrows=20)
    for i in range(len(preview)):
        vals = [str(v).strip() if pd.notna(v) else '' for v in preview.iloc[i].tolist()]
        if '종목코드' in vals and '종목명' in vals:
            return i
    raise RuntimeError('TSV에서 종목코드/종목명 헤더 행을 찾지 못했습니다.')


def find_main_sheet_and_header(path: Path) -> Tuple[str, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if '종목코드' in vals and '종목명' in vals:
                return ws.title, ridx
    raise RuntimeError('종목코드/종목명이 있는 헤더 행을 찾지 못했습니다.')


def load_main_table(path: Path) -> Tuple[pd.DataFrame, str, int]:
    ext = path.suffix.lower()
    if ext == '.tsv':
        header_row0 = find_tsv_header_row(path)
        df = pd.read_csv(path, sep='\t', header=header_row0, dtype=object)
        df = df.loc[:, ~df.columns.astype(str).str.contains(r'^Unnamed')]
        if '종목코드' not in df.columns or '종목명' not in df.columns:
            raise RuntimeError('메인 TSV에서 종목코드/종목명 컬럼을 찾지 못했습니다.')
        df['종목코드'] = df['종목코드'].map(norm_code)
        df = df[df['종목코드'] != ''].copy()
        return df, 'TSV', header_row0 + 1

    sheet_name, header_row = find_main_sheet_and_header(path)
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, dtype=object)
    df = df.loc[:, ~df.columns.astype(str).str.contains(r'^Unnamed')]
    if '종목코드' not in df.columns or '종목명' not in df.columns:
        raise RuntimeError('메인 시트에서 종목코드/종목명 컬럼을 찾지 못했습니다.')
    df['종목코드'] = df['종목코드'].map(norm_code)
    df = df[df['종목코드'] != ''].copy()
    return df, sheet_name, header_row


def safe_float(x) -> float:
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return np.nan
        s = str(x).replace(',', '').strip()
        if s == '':
            return np.nan
        return float(s)
    except Exception:
        return np.nan


def import_pykrx():
    try:
        from pykrx import stock  # type: ignore
    except Exception as e:
        raise RuntimeError(
            'pykrx가 필요합니다. 먼저 `pip install pykrx pandas openpyxl numpy` 실행하세요.'
        ) from e
    return stock


def get_market_type_map(stock, codes: List[str]) -> Dict[str, str]:
    today = datetime.today().strftime('%Y%m%d')
    market_map: Dict[str, str] = {}
    for market in ('KOSPI', 'KOSDAQ', 'KONEX'):
        try:
            tickers = set(stock.get_market_ticker_list(today, market=market))
        except TypeError:
            tickers = set(stock.get_market_ticker_list(market=market))
        for code in codes:
            if code in tickers:
                market_map[code] = market
    return market_map


def get_benchmark_index_ohlcv(stock, market: str, start: str, end: str) -> pd.DataFrame:
    # pykrx index ticker는 환경별로 이름이 다를 수 있어 유연하게 찾는다.
    target_names = {
        'KOSPI': ['코스피', 'KOSPI'],
        'KOSDAQ': ['코스닥', 'KOSDAQ'],
    }.get(market, [])
    for m in ('KOSPI', 'KOSDAQ'):
        try:
            tickers = stock.get_index_ticker_list(market=m)
        except Exception:
            continue
        for ticker in tickers:
            try:
                name = stock.get_index_ticker_name(ticker)
            except Exception:
                name = ''
            if name in target_names:
                try:
                    df = stock.get_index_ohlcv_by_date(start, end, ticker)
                    if not df.empty:
                        return df
                except Exception:
                    pass
    return pd.DataFrame()


def fetch_ohlcv(stock, code: str, start: str, end: str) -> pd.DataFrame:
    df = stock.get_market_ohlcv_by_date(start, end, code, adjusted=True)
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    rename_map = {
        '시가': 'Open',
        '고가': 'High',
        '저가': 'Low',
        '종가': 'Close',
        '거래량': 'Volume',
        '거래대금': 'Value',
        '등락률': 'Return',
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    return out


def compute_ma(series: pd.Series, win: int) -> pd.Series:
    return series.rolling(win).mean()


def find_pivot_highs(close: pd.Series, window: int = 3) -> List[Tuple[pd.Timestamp, float]]:
    vals = close.astype(float).values
    idx = close.index
    pivots = []
    if len(vals) < window * 2 + 1:
        return pivots
    for i in range(window, len(vals) - window):
        left = vals[i - window:i]
        right = vals[i + 1:i + 1 + window]
        c = vals[i]
        if np.isnan(c):
            continue
        if c >= np.nanmax(left) and c >= np.nanmax(right):
            pivots.append((idx[i], float(c)))
    return pivots


def stock_timing_metrics(df: pd.DataFrame, pivot_window: int, breakout_lookback: int, late_entry_pct: float) -> Dict[str, object]:
    if df.empty or 'Close' not in df.columns or 'Volume' not in df.columns:
        return {}

    close = df['Close'].astype(float)
    vol = df['Volume'].astype(float)
    ma30 = compute_ma(close, 30)
    last_close = float(close.iloc[-1])
    last_ma30 = float(ma30.iloc[-1]) if not pd.isna(ma30.iloc[-1]) else np.nan
    ma30_prev5 = float(ma30.iloc[-6]) if len(ma30) >= 6 and not pd.isna(ma30.iloc[-6]) else np.nan

    ma30_above = bool(not pd.isna(last_ma30) and last_close > last_ma30)
    ma30_up = bool(not pd.isna(last_ma30) and not pd.isna(ma30_prev5) and last_ma30 > ma30_prev5)

    hist = close.tail(breakout_lookback)
    pivots = find_pivot_highs(hist, window=pivot_window)

    # 최근 5일은 돌파 확인용이므로 저항 후보에서 제외
    cutoff_date = hist.index[-5] if len(hist) >= 5 else hist.index[-1]
    pivot_prices = [p for dt, p in pivots if dt < cutoff_date]

    overhead = sorted([p for p in pivot_prices if p > last_close])
    nearest_overhead = overhead[0] if overhead else np.nan
    overhead_margin_pct = ((nearest_overhead / last_close) - 1.0) * 100.0 if overhead else np.nan
    overhead_count_10 = sum(1 for p in overhead if p <= last_close * 1.10)
    overhead_light = overhead_count_10 <= 1

    breakout_ref = np.nan
    if pivot_prices:
        breakout_ref = max(pivot_prices)
    elif len(hist) > 20:
        breakout_ref = float(hist.iloc[:-5].max()) if len(hist.iloc[:-5]) else np.nan

    breakout = bool(not pd.isna(breakout_ref) and last_close > breakout_ref)
    late_distance_pct = ((last_close / breakout_ref) - 1.0) * 100.0 if breakout and not pd.isna(breakout_ref) else np.nan
    not_late = bool(breakout and not pd.isna(late_distance_pct) and late_distance_pct <= late_entry_pct)

    prev20_avg_vol = float(vol.iloc[-25:-5].mean()) if len(vol) >= 25 else np.nan
    last5_max_vol = float(vol.iloc[-5:].max()) if len(vol) >= 5 else np.nan
    ratio_5d = (last5_max_vol / prev20_avg_vol) if (not pd.isna(last5_max_vol) and not pd.isna(prev20_avg_vol) and prev20_avg_vol > 0) else np.nan

    last20_sum = float(vol.iloc[-20:].sum()) if len(vol) >= 20 else np.nan
    prev20_sum = float(vol.iloc[-40:-20].sum()) if len(vol) >= 40 else np.nan
    ratio_20d = (last20_sum / prev20_sum) if (not pd.isna(last20_sum) and not pd.isna(prev20_sum) and prev20_sum > 0) else np.nan

    vol_surge = bool((not pd.isna(ratio_5d) and ratio_5d >= 2.0) or (not pd.isna(ratio_20d) and ratio_20d >= 2.0))

    return {
        '종가': last_close,
        'MA30': last_ma30,
        'MA30위': 'Y' if ma30_above else 'N',
        'MA30우상향': 'Y' if ma30_up else 'N',
        '최근피벗고점개수': len(pivot_prices),
        '윗저항개수(10%)': overhead_count_10,
        '가장가까운윗저항': nearest_overhead,
        '윗저항여유(%)': overhead_margin_pct,
        '윗저항적음': 'Y' if overhead_light else 'N',
        '돌파기준가격': breakout_ref,
        '저항돌파': 'Y' if breakout else 'N',
        '1주최고거래량/지난달평균': ratio_5d,
        '최근20일누적/직전20일누적': ratio_20d,
        '돌파거래량급증': 'Y' if vol_surge else 'N',
        '늦은진입거리(%)': late_distance_pct,
        '늦은진입아님': 'Y' if not_late else 'N',
    }


def market_trend_metrics(df: pd.DataFrame) -> Dict[str, object]:
    if df.empty or '종가' not in df.columns:
        return {'시장상승세': '보류', '시장사유': '지수 데이터 없음'}
    close = df['종가'].astype(float)
    ma30 = close.rolling(30).mean()
    if len(close) < 35 or pd.isna(ma30.iloc[-1]) or pd.isna(ma30.iloc[-6]):
        return {'시장상승세': '보류', '시장사유': '지수 데이터 부족'}
    above = close.iloc[-1] > ma30.iloc[-1]
    up = ma30.iloc[-1] > ma30.iloc[-6]
    if above and up:
        return {'시장상승세': 'Y', '시장사유': '지수>MA30 & MA30우상향'}
    if above or up:
        return {'시장상승세': '보류', '시장사유': '절반만 충족'}
    return {'시장상승세': 'N', '시장사유': '지수<MA30 or MA30하락'}


def build_group_baskets(ohlcv_map: Dict[str, pd.DataFrame], base_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    baskets: Dict[str, pd.DataFrame] = {}
    if '그룹' not in base_df.columns:
        return baskets
    group_map: Dict[str, List[str]] = {}
    for _, row in base_df.iterrows():
        grp = str(row.get('그룹', '')).strip()
        code = norm_code(row.get('종목코드', ''))
        if grp and code in ohlcv_map and not ohlcv_map[code].empty:
            group_map.setdefault(grp, []).append(code)
    for grp, codes in group_map.items():
        aligned = []
        for code in codes:
            s = ohlcv_map[code]['Close'].astype(float).rename(code)
            s = s / s.iloc[0] if len(s) and s.iloc[0] != 0 else s
            aligned.append(s)
        if not aligned:
            continue
        df = pd.concat(aligned, axis=1).dropna(how='all')
        if df.empty:
            continue
        baskets[grp] = pd.DataFrame({'Close': df.mean(axis=1)})
    return baskets


def group_trend_label(basket_df: Optional[pd.DataFrame]) -> Tuple[str, str]:
    if basket_df is None or basket_df.empty:
        return '보류', '그룹 바스켓 데이터 없음'
    close = basket_df['Close'].astype(float)
    ma30 = close.rolling(30).mean()
    if len(close) < 35 or pd.isna(ma30.iloc[-1]) or pd.isna(ma30.iloc[-6]):
        return '보류', '그룹 데이터 부족'
    above = close.iloc[-1] > ma30.iloc[-1]
    up = ma30.iloc[-1] > ma30.iloc[-6]
    if above and up:
        return 'Y', '그룹바스켓>MA30 & MA30우상향'
    if above or up:
        return '보류', '그룹 절반만 충족'
    return 'N', '그룹바스켓<MA30 or MA30하락'


def score_and_label(row: pd.Series) -> Tuple[int, str, str]:
    score = 0
    reasons = []
    if row.get('시장상승세') == 'Y':
        score += 1
    else:
        reasons.append('시장상승세 미확인')
    if row.get('업종상승세') == 'Y':
        score += 1
    else:
        reasons.append('업종상승세 미확인')
    if row.get('MA30위') == 'Y':
        score += 1
    else:
        reasons.append('MA30 아래')
    if row.get('MA30우상향') == 'Y':
        score += 1
    else:
        reasons.append('MA30 우하향/평탄')
    if row.get('윗저항적음') == 'Y':
        score += 1
    else:
        reasons.append('윗저항 많음')
    if row.get('저항돌파') == 'Y':
        score += 1
    else:
        reasons.append('저항 미돌파')
    if row.get('돌파거래량급증') == 'Y':
        score += 1
    else:
        reasons.append('거래량 급증 미확인')
    if row.get('늦은진입아님') == 'Y':
        score += 1
    else:
        reasons.append('늦은 진입 또는 돌파 없음')

    if row.get('저항돌파') == 'Y' and row.get('돌파거래량급증') == 'Y' and row.get('늦은진입아님') == 'Y' and row.get('시장상승세') == 'Y' and row.get('업종상승세') == 'Y':
        label = 'A. 매수검토'
    elif score >= 5:
        label = 'B. 관찰'
    elif score >= 3:
        label = 'C. 보류'
    else:
        label = 'D. 제외/후순위'

    return score, label, '; '.join(reasons[:5])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', '--in', dest='input', required=True)
    ap.add_argument('--output', '--out', dest='output', required=True)
    ap.add_argument('--pivot-window', type=int, default=3)
    ap.add_argument('--breakout-lookback', type=int, default=120)
    ap.add_argument('--late-entry-pct', type=float, default=5.0)
    ap.add_argument('--lookback-days', type=int, default=420)
    args = ap.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise SystemExit(f'input not found: {input_path}')

    df, main_sheet, header_row = load_main_table(input_path)
    print(f'main_sheet={main_sheet} header_row={header_row} rows={len(df)}')

    stock = import_pykrx()
    end_dt = datetime.today()
    start_dt = end_dt - timedelta(days=args.lookback_days)
    start = start_dt.strftime('%Y%m%d')
    end = end_dt.strftime('%Y%m%d')

    codes = [norm_code(x) for x in df['종목코드'].tolist()]
    market_type_map = get_market_type_map(stock, codes)

    # 종목별 가격/거래량 로드
    ohlcv_map: Dict[str, pd.DataFrame] = {}
    for code in codes:
        try:
            ohlcv_map[code] = fetch_ohlcv(stock, code, start, end)
        except Exception:
            ohlcv_map[code] = pd.DataFrame()

    # 시장 지수 로드
    market_frames: Dict[str, pd.DataFrame] = {}
    market_summary_rows = []
    for market in ('KOSPI', 'KOSDAQ'):
        try:
            idx_df = get_benchmark_index_ohlcv(stock, market, start, end)
        except Exception:
            idx_df = pd.DataFrame()
        market_frames[market] = idx_df
        mm = market_trend_metrics(idx_df)
        market_summary_rows.append({'시장': market, **mm})

    market_summary = pd.DataFrame(market_summary_rows)

    # 그룹 바스켓
    group_baskets = build_group_baskets(ohlcv_map, df)

    out_rows = []
    base_cols = [
        '종목코드', '종목명', '그룹', '판정구분', '현재가', '하드필터통과', '종합판정',
        '린치PER판정', '린치PER배수', '연간이익증가율(3년CAGR,%)', '그레이엄괴리율(3년,%)',
    ]
    existing_cols = [c for c in base_cols if c in df.columns]

    for _, row in df.iterrows():
        code = norm_code(row.get('종목코드', ''))
        group = str(row.get('그룹', '')).strip()
        mkt = market_type_map.get(code, 'KOSPI')
        market_state = market_trend_metrics(market_frames.get(mkt, pd.DataFrame()))
        group_state, group_reason = group_trend_label(group_baskets.get(group))
        timing = stock_timing_metrics(
            ohlcv_map.get(code, pd.DataFrame()),
            pivot_window=args.pivot_window,
            breakout_lookback=args.breakout_lookback,
            late_entry_pct=args.late_entry_pct,
        )
        out = {c: row.get(c, '') for c in existing_cols}
        out['시장구분'] = mkt
        out['시장상승세'] = market_state.get('시장상승세', '보류')
        out['시장사유'] = market_state.get('시장사유', '')
        out['업종상승세'] = group_state
        out['업종사유'] = group_reason
        out.update(timing)
        score, label, reason = score_and_label(pd.Series(out))
        out['기술점수'] = score
        out['타이밍종합판정'] = label
        out['타이밍사유'] = reason
        out_rows.append(out)

    timing_df = pd.DataFrame(out_rows)

    if output_path.suffix.lower() == '.tsv':
        merged_df = df.copy()
        join_cols = ['종목코드']
        if '종목명' in merged_df.columns and '종목명' in timing_df.columns:
            join_cols.append('종목명')
        timing_only_cols = [c for c in timing_df.columns if c not in merged_df.columns]
        timing_export_df = timing_df[join_cols + timing_only_cols].copy()
        merged_df = merged_df.merge(timing_export_df, on=join_cols, how='left')
        merged_df.to_csv(output_path, sep='	', index=False)
        print(f'saved: {output_path}')
        return

    shutil.copy2(input_path, output_path)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        skip_sheets = {'시장자동체크', '타이밍자동체크', '책기준_타이밍'}

        xls = pd.ExcelFile(input_path, engine='openpyxl')

        for s in xls.sheet_names:
            if s in skip_sheets:
                continue
            pd.read_excel(input_path, sheet_name=s, header=None, dtype=object).to_excel(
                writer, sheet_name=s, header=False, index=False
            )

        market_summary.to_excel(writer, sheet_name='시장자동체크', index=False)
        timing_df.to_excel(writer, sheet_name='타이밍자동체크', index=False)

    print(f'saved: {output_path}')


if __name__ == '__main__':
    main()
